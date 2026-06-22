import os
import sys
from typing import List
from PySide6.QtWidgets import (
    QMainWindow, QWidget, QHBoxLayout, QVBoxLayout, QLabel, QPushButton,
    QStatusBar, QMessageBox, QButtonGroup, QFileDialog, QStackedWidget,
    QFrame, QSizePolicy
)
from PySide6.QtCore import Qt, QSize
from PySide6.QtGui import QIcon, QFont, QAction, QKeySequence

from .widgets import get_app_style
from .import_page import ImportPage
from .validation_page import ValidationPage
from .batch_page import BatchAdjustPage
from .print_page import PrintPage
from .backup_page import BackupPage
from ..core.models import PriceItem
from ..core.data_manager import DataStore


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        data_dir = os.path.join(base_dir, "data")
        backup_dir = os.path.join(base_dir, "backups")
        self.data_store = DataStore(data_dir, backup_dir)

        self.setWindowTitle("医美价目表批量维护工具")
        self.setMinimumSize(1280, 800)
        self.resize(1440, 900)
        self.setStyleSheet(get_app_style())

        self._init_ui()
        self._connect_signals()
        self._sync_items_from_store()

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QHBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        sidebar = QFrame()
        sidebar.setObjectName("Sidebar")
        sidebar.setFixedWidth(220)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(0)

        title = QLabel("医美价目管家")
        title.setObjectName("SidebarTitle")
        sidebar_layout.addWidget(title)

        subtitle = QLabel("Price Management")
        subtitle.setStyleSheet("""
            color: #64748b; padding: 0px 20px 15px 20px;
            font-size: 9pt; border-bottom: 1px solid #334155;
        """)
        sidebar_layout.addWidget(subtitle)

        self.nav_buttons = []
        nav_config = [
            ("📥  导入表格", 0, "导入 Excel 价目表"),
            ("✅  价格校验", 1, "检查价格合理性"),
            ("⚡  批量调整", 2, "批量调整价格和名称"),
            ("🖨️  打印预览", 3, "导出和打印价目表"),
            ("🔙  备份恢复", 4, "快照管理与回滚"),
        ]
        for text, idx, tip in nav_config:
            btn = QPushButton(text)
            btn.setObjectName("NavButton")
            btn.setCheckable(True)
            btn.setToolTip(tip)
            btn.setCursor(Qt.PointingHandCursor)
            if idx == 0:
                btn.setChecked(True)
            sidebar_layout.addWidget(btn)
            self.nav_buttons.append((btn, idx))

        sidebar_layout.addStretch()

        save_section = QFrame()
        save_section.setStyleSheet("padding: 12px 15px; border-top: 1px solid #334155;")
        save_layout = QVBoxLayout(save_section)
        save_layout.setContentsMargins(0, 0, 0, 0)
        save_layout.setSpacing(8)

        self.btn_save_all = QPushButton("💾 保存全部修改")
        self.btn_save_all.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 10px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #2563eb; }
        """)
        save_layout.addWidget(self.btn_save_all)

        self.btn_export_sample = QPushButton("📄 导出示例模板")
        self.btn_export_sample.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #cbd5e1;
                border: 1px solid #475569;
                border-radius: 6px;
                padding: 8px;
                font-size: 9pt;
            }
            QPushButton:hover {
                background-color: #334155;
                color: white;
            }
        """)
        save_layout.addWidget(self.btn_export_sample)

        sidebar_layout.addWidget(save_section)

        version_label = QLabel("v1.0.0 · 离线版")
        version_label.setStyleSheet("""
            color: #475569; padding: 10px 20px; font-size: 8pt;
            border-top: 1px solid #334155;
        """)
        sidebar_layout.addWidget(version_label)

        main_layout.addWidget(sidebar)

        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(0, 0, 0, 0)

        topbar = QFrame()
        topbar.setStyleSheet("""
            QFrame {
                background-color: white;
                border-bottom: 1px solid #e2e8f0;
                padding: 10px 25px;
            }
        """)
        topbar_layout = QHBoxLayout(topbar)
        topbar_layout.setContentsMargins(0, 0, 0, 0)

        self.topbar_title = QLabel("导入表格")
        self.topbar_title.setStyleSheet("font-size: 13pt; font-weight: bold; color: #1e293b;")
        topbar_layout.addWidget(self.topbar_title)

        topbar_layout.addStretch()

        self.item_count_label = QLabel("项目数：0")
        self.item_count_label.setStyleSheet("""
            padding: 6px 14px; background: #eff6ff; color: #1e40af;
            border-radius: 6px; font-weight: 600;
        """)
        topbar_layout.addWidget(self.item_count_label)

        content_layout.addWidget(topbar)

        self.stack = QStackedWidget()
        self.import_page = ImportPage()
        self.validation_page = ValidationPage()
        self.batch_page = BatchAdjustPage()
        self.print_page = PrintPage()
        self.backup_page = BackupPage(backup_dir=self.data_store.backup_manager.backup_dir)

        self.stack.addWidget(self.import_page)
        self.stack.addWidget(self.validation_page)
        self.stack.addWidget(self.batch_page)
        self.stack.addWidget(self.print_page)
        self.stack.addWidget(self.backup_page)

        content_layout.addWidget(self.stack, 1)
        main_layout.addWidget(content, 1)

        self.statusBar().showMessage("就绪")
        self.statusBar().setStyleSheet("padding: 6px 20px;")

    def _connect_signals(self):
        self.nav_group = QButtonGroup(self)
        self.nav_group.setExclusive(True)
        for btn, idx in self.nav_buttons:
            self.nav_group.addButton(btn, idx)
            btn.clicked.connect(lambda checked, i=idx: self._switch_page(i))

        self.import_page.itemsImported.connect(self._on_items_imported)
        self.validation_page.itemsUpdated.connect(self._on_items_updated)
        self.batch_page.itemsUpdated.connect(self._on_items_updated)
        self.backup_page.rollbackRequested.connect(self._on_rollback)

        self.btn_save_all.clicked.connect(self._save_all)
        self.btn_export_sample.clicked.connect(self._export_sample_template)

    def _switch_page(self, idx: int):
        self.stack.setCurrentIndex(idx)
        titles = ["导入表格", "价格校验", "批量调整", "打印预览", "备份恢复"]
        self.topbar_title.setText(titles[idx])

        if idx == 1:
            self.validation_page.set_items(self.data_store.items)
        elif idx == 2:
            self.batch_page.set_items(self.data_store.items)
        elif idx == 3:
            self.print_page.set_items(self.data_store.items)
        elif idx == 4:
            self.backup_page.set_items(self.data_store.items)

    def _sync_items_from_store(self):
        self.item_count_label.setText(f"项目数：{len(self.data_store.items)}")

    def _on_items_imported(self, items: List[PriceItem]):
        self.data_store.set_items(items)
        try:
            ok, errors = self.data_store.save(
                create_snapshot=True,
                snapshot_name=f"导入_{len(items)}项",
                snapshot_desc=f"从外部文件导入的初始价目表数据"
            )
            if errors:
                QMessageBox.warning(self, "部分校验问题",
                                    f"已保存，但有 {len(errors)} 个项目存在校验问题，建议前往「价格校验」页面处理。")
        except Exception as e:
            QMessageBox.warning(self, "保存提示",
                                f"数据已加载，保存快照时提示：{str(e)}。\n可手动到备份页面创建快照。")
            self.data_store.items = items

        self._sync_items_from_store()
        self.statusBar().showMessage(f"已导入 {len(items)} 个项目", 5000)

        self.nav_buttons[1][0].setChecked(True)
        self._switch_page(1)

    def _on_items_updated(self, items: List[PriceItem]):
        self.data_store.set_items(items)
        self._sync_items_from_store()
        self.statusBar().showMessage(f"已更新 {len(items)} 个项目", 3000)

    def _on_rollback(self, items: List[PriceItem]):
        self.data_store.set_items(items)
        try:
            self.data_store._save_data_only() if hasattr(self.data_store, "_save_data_only") else None
        except Exception:
            pass
        self._sync_items_from_store()
        self.statusBar().showMessage(f"已回滚到 {len(items)} 个项目", 3000)

        for page_idx in [1, 2, 3]:
            if page_idx == self.stack.currentIndex():
                self._switch_page(page_idx)

    def _save_all(self):
        if not self.data_store.items:
            QMessageBox.information(self, "提示", "当前没有项目数据，无需保存")
            return

        ok, errors = self.data_store.save(
            create_snapshot=True,
            snapshot_name=f"手动保存_{len(self.data_store.items)}项",
            snapshot_desc="用户点击'保存全部修改'触发"
        )

        if ok:
            self._sync_items_from_store()
            QMessageBox.information(self, "保存成功",
                                    f"已保存 {len(self.data_store.items)} 个项目\n"
                                    f"并创建了对应快照备份。")
            self.statusBar().showMessage("保存成功", 3000)
        else:
            reply = QMessageBox.question(
                self, "保存提示",
                f"存在 {len(errors)} 个校验问题，部分操作可能无法正常工作。\n\n"
                + "\n".join(errors[:5]) + ("\n..." if len(errors) > 5 else "") +
                "\n\n是否仍要强制保存？",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                import json
                with open(self.data_store.data_file, "w", encoding="utf-8") as f:
                    json.dump([it.to_dict() for it in self.data_store.items], f, ensure_ascii=False, indent=2)
                self._sync_items_from_store()
                QMessageBox.information(self, "已保存", "数据已保存，但未创建快照。")

    def _export_sample_template(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出示例模板", "医美价目表_模板.xlsx", "Excel 文件 (*.xlsx)"
        )
        if not file_path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "价目表"

        headers = ["项目名称", "前台展示名", "内部核算名", "分类",
                  "原价", "会员价", "医生费", "耗材费", "耗材品牌",
                  "成本价", "备注"]
        sample_data = [
            ["热玛吉四代 900发", "热玛吉四代", "TMJ-S-0900", "光电类",
             12800, 9800, 3000, 1500, "Thermage", 6000, "含下颌缘提升"],
            ["乔雅登雅致 0.8ml", "乔雅登雅致填充", "QYD-Y-08", "注射类",
             6800, 5800, 800, 3500, "乔雅登", 4500, "单支价格 不稀释"],
            ["超微小气泡 清洁补水", "小气泡清洁", "XQP-C-01", "清洁补水",
             298, 99, 50, 20, "—", 70, "首次体验价"],
            ["切开双眼皮", "全切双眼皮", "SSS-Y-01", "手术类",
             6800, 4800, 2500, 500, "—", 3000, "含局麻 拆线"],
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for r, row_data in enumerate(sample_data, 2):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if c >= 5 and isinstance(val, (int, float)):
                    cell.number_format = "¥#,##0"

        note_row = len(sample_data) + 3
        ws.cell(row=note_row, column=1, value="📝 使用说明：").font = Font(bold=True)
        notes = [
            "1. 项目名称、分类、原价为必填项",
            "2. 分类可选：光电类 / 注射类 / 手术类 / 清洁补水 / 皮肤护理 / 其他",
            "3. 所有价格为数字，无需输入¥单位",
            "4. 保存后可直接拖入本工具导入窗口使用"
        ]
        for i, note in enumerate(notes, 1):
            ws.cell(row=note_row + i, column=1, value=note).font = Font(italic=True, color="666666")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 16
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["K"].width = 24

        try:
            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"模板已保存到：\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出失败：{str(e)}")
