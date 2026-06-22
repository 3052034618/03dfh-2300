import os
import json
import shutil
from typing import List, Optional, Dict, Any, Tuple
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import PriceItem, Snapshot, CategoryManager
from .price_engine import PriceValidator


class ExcelImporter:
    COLUMN_MAPPING = {
        "name": ["项目名称", "名称", "项目", "品名", "产品名称", "name", "item"],
        "display_name": ["前台展示名", "展示名", "显示名称", "display"],
        "internal_name": ["内部核算名", "内部名称", "核算名", "internal"],
        "category": ["分类", "类别", "项目分类", "大类", "category", "type"],
        "original_price": ["原价", "标准价", "门市价", "零售价", "价格", "original", "price"],
        "member_price": ["会员价", "会员价格", "优惠价", "折扣价", "member", "vip"],
        "doctor_fee": ["医生费", "医师费", "操作费", "手工费", "doctor"],
        "material_fee": ["耗材费", "材料费", "产品费", "material", "supply"],
        "material_brand": ["耗材品牌", "品牌", "厂家", "brand"],
        "cost_price": ["成本价", "底价", "进货价", "cost"],
        "remark": ["备注", "说明", "注意事项", "remark", "note"]
    }

    @staticmethod
    def _detect_column(header: str, field_keys: List[str]) -> Optional[str]:
        header_lower = str(header).strip().lower()
        for key in field_keys:
            if key.lower() in header_lower or header_lower in key.lower():
                return True
        return False

    @classmethod
    def _map_columns(cls, headers: List[str]) -> Dict[str, int]:
        mapping = {}
        for idx, header in enumerate(headers):
            header_str = str(header).strip()
            for field_name, keywords in cls.COLUMN_MAPPING.items():
                if field_name in mapping:
                    continue
                if cls._detect_column(header_str, keywords):
                    mapping[field_name] = idx
                    break
        return mapping

    @staticmethod
    def _safe_float(value: Any) -> float:
        if value is None or value == "":
            return 0.0
        try:
            if isinstance(value, str):
                cleaned = value.replace(",", "").replace("¥", "").replace("￥", "").replace("元", "").strip()
                return round(float(cleaned), 2) if cleaned else 0.0
            return round(float(value), 2)
        except (ValueError, TypeError):
            return 0.0

    @classmethod
    def inspect_file(cls, file_path: str, preview_rows: int = 5) -> Tuple[List[str], List[List[Any]], List[str]]:
        """仅读取表头和前 N 行预览，不做字段映射，供用户手动匹配列。
        返回 (headers, preview_rows, warnings)
        """
        warnings = []
        if not os.path.exists(file_path):
            return [], [], [f"文件不存在: {file_path}"]
        try:
            ext = os.path.splitext(file_path)[1].lower()
            if ext in [".xlsx", ".xls"]:
                df = pd.read_excel(file_path, dtype=object)
            elif ext == ".csv":
                df = pd.read_csv(file_path, dtype=object)
            else:
                return [], [], [f"不支持的文件格式: {ext}"]
            if df.empty:
                return [], [], ["表格内容为空"]
            headers = [str(h).strip() if h is not None else f"列{i+1}" for i, h in enumerate(df.columns)]
            preview = []
            for _, row in df.head(preview_rows).iterrows():
                preview.append([("" if v is None else str(v)) for v in row.tolist()])
            # 补全：如果有的行比 header 短，用空串填充
            for r in preview:
                while len(r) < len(headers):
                    r.append("")
            return headers, preview, warnings
        except Exception as e:
            return [], [], [f"读取文件失败: {str(e)}"]

    @classmethod
    def import_file(cls, file_path: str,
                    col_mapping: Optional[Dict[str, int]] = None) -> Tuple[List[PriceItem], List[str]]:
        """导入 Excel。如果传了 col_mapping 则使用用户手动匹配的结果，否则自动识别。"""
        if not os.path.exists(file_path):
            return [], [f"文件不存在: {file_path}"]

        warnings = []
        items = []

        try:
            ext = os.path.splitext(file_path)[1].lower()
            if ext in [".xlsx", ".xls"]:
                df = pd.read_excel(file_path, dtype=object)
            elif ext == ".csv":
                df = pd.read_csv(file_path, dtype=object)
            else:
                return [], [f"不支持的文件格式: {ext}"]

            if df.empty:
                return [], ["表格内容为空"]

            headers = list(df.columns)
            if col_mapping is not None and len(col_mapping) > 0:
                # 使用用户手动匹配
                col_mapping = {k: v for k, v in col_mapping.items() if v is not None and v >= 0}
            else:
                col_mapping = cls._map_columns(headers)

            if "name" not in col_mapping and len(headers) > 0:
                warnings.append("未识别到'项目名称'列，将尝试使用第一列作为名称")
                col_mapping["name"] = 0

            for _, row in df.iterrows():
                values = row.tolist()
                item = PriceItem()

                for field_name, col_idx in col_mapping.items():
                    if col_idx >= len(values):
                        continue
                    value = values[col_idx]

                    if field_name in ["original_price", "member_price", "doctor_fee",
                                     "material_fee", "cost_price"]:
                        setattr(item, field_name, cls._safe_float(value))
                    elif field_name == "category":
                        cat = str(value).strip() if value else ""
                        if not cat:
                            name_val = values[col_mapping["name"]] if "name" in col_mapping else ""
                            cat = CategoryManager.detect_category(str(name_val))
                        item.category = cat
                    else:
                        setattr(item, field_name, str(value).strip() if value else "")

                if not item.name:
                    continue

                if not item.category:
                    item.category = CategoryManager.detect_category(item.name)

                if not item.display_name:
                    item.display_name = item.name

                items.append(item)

        except Exception as e:
            return [], [f"导入失败: {str(e)}"]

        return items, warnings

    @staticmethod
    def export_items(items: List[PriceItem], file_path: str,
                     include_internal: bool = True) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "价目表"

        if include_internal:
            headers = ["项目名称", "前台展示名", "内部核算名", "分类",
                      "原价", "会员价", "医生费", "耗材费", "耗材品牌",
                      "成本价", "总成本", "备注"]
        else:
            headers = ["项目名称", "分类", "原价", "会员价", "备注"]

        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, item in enumerate(items, 2):
            if include_internal:
                row_data = [
                    item.name, item.display_name, item.internal_name, item.category,
                    item.original_price, item.member_price,
                    item.doctor_fee, item.material_fee, item.material_brand,
                    item.cost_price, item.total_cost, item.remark
                ]
            else:
                row_data = [
                    item.display_name or item.name, item.category,
                    item.original_price, item.member_price, item.remark
                ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")
                if isinstance(value, float) and col >= 5 and include_internal:
                    cell.number_format = "¥#,##0.00"
                elif isinstance(value, float) and not include_internal and col >= 3:
                    cell.number_format = "¥#,##0.00"

                if item.is_conflict:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        for col in range(1, len(headers) + 1):
            max_length = len(str(headers[col - 1]))
            for row in range(2, len(items) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_length + 4, 30)

        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        wb.save(file_path)
        return file_path


class BackupManager:
    def __init__(self, backup_dir: str):
        self.backup_dir = backup_dir
        os.makedirs(self.backup_dir, exist_ok=True)
        self.index_file = os.path.join(self.backup_dir, "index.json")
        self._load_index()

    def _load_index(self):
        if os.path.exists(self.index_file):
            with open(self.index_file, "r", encoding="utf-8") as f:
                self.index = json.load(f)
        else:
            self.index = {"snapshots": []}

    def _save_index(self):
        with open(self.index_file, "w", encoding="utf-8") as f:
            json.dump(self.index, f, ensure_ascii=False, indent=2)

    def create_snapshot(self, items: List[PriceItem], name: str = "",
                       description: str = "") -> Snapshot:
        snapshot = Snapshot(
            name=name or f"快照_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
            description=description,
            items=[item.to_dict() for item in items],
            item_count=len(items)
        )

        snapshot_file = os.path.join(self.backup_dir, f"{snapshot.id}.json")
        with open(snapshot_file, "w", encoding="utf-8") as f:
            json.dump(snapshot.to_dict(), f, ensure_ascii=False, indent=2)

        self.index["snapshots"].insert(0, {
            "id": snapshot.id,
            "name": snapshot.name,
            "description": snapshot.description,
            "created_at": snapshot.created_at,
            "item_count": snapshot.item_count,
            "file": snapshot_file
        })
        self._save_index()

        return snapshot

    def list_snapshots(self) -> List[Dict[str, Any]]:
        return self.index.get("snapshots", [])

    def load_snapshot(self, snapshot_id: str) -> Optional[List[PriceItem]]:
        snapshot_file = os.path.join(self.backup_dir, f"{snapshot_id}.json")
        if not os.path.exists(snapshot_file):
            return None

        try:
            with open(snapshot_file, "r", encoding="utf-8") as f:
                data = json.load(f)

            items = []
            for item_data in data.get("items", []):
                item = PriceItem(
                    id=item_data.get("id", ""),
                    name=item_data.get("name", ""),
                    display_name=item_data.get("display_name", ""),
                    internal_name=item_data.get("internal_name", ""),
                    category=item_data.get("category", ""),
                    original_price=item_data.get("original_price", 0.0),
                    member_price=item_data.get("member_price", 0.0),
                    doctor_fee=item_data.get("doctor_fee", 0.0),
                    material_fee=item_data.get("material_fee", 0.0),
                    material_brand=item_data.get("material_brand", ""),
                    cost_price=item_data.get("cost_price", 0.0),
                    remark=item_data.get("remark", ""),
                    is_conflict=item_data.get("is_conflict", False),
                    conflict_ids=item_data.get("conflict_ids", []),
                    created_at=item_data.get("created_at", ""),
                    updated_at=item_data.get("updated_at", "")
                )
                items.append(item)
            return items
        except Exception:
            return None

    def delete_snapshot(self, snapshot_id: str) -> bool:
        snapshot_file = os.path.join(self.backup_dir, f"{snapshot_id}.json")
        if os.path.exists(snapshot_file):
            os.remove(snapshot_file)

        self.index["snapshots"] = [
            s for s in self.index.get("snapshots", []) if s.get("id") != snapshot_id
        ]
        self._save_index()
        return True

    def rollback_to(self, snapshot_id: str) -> Optional[List[PriceItem]]:
        items = self.load_snapshot(snapshot_id)
        if items is not None:
            self.create_snapshot(items, name=f"回滚前备份_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        return items


class MappingTemplateManager:
    """字段映射模板管理：按门店保存列映射规则，下次自动套用。

    模板数据结构：
    {
        "store_name": "朝阳门店",
        "headers": ["项目名称", "原价", ...],  # 保存时的表头列表（用于匹配度计算）
        "mapping": {"name": 0, "original_price": 2, ...},  # 字段 -> 列索引
        "created_at": "2025-01-01T...",
        "last_used_at": "2025-01-02T...",
        "use_count": 3
    }
    """

    def __init__(self, data_dir: str):
        self.data_dir = data_dir
        os.makedirs(self.data_dir, exist_ok=True)
        self.template_file = os.path.join(self.data_dir, "mapping_templates.json")
        self.templates: List[Dict[str, Any]] = []
        self._load()

    def _load(self):
        if os.path.exists(self.template_file):
            try:
                with open(self.template_file, "r", encoding="utf-8") as f:
                    self.templates = json.load(f)
            except Exception:
                self.templates = []
        else:
            self.templates = []

    def _save(self):
        with open(self.template_file, "w", encoding="utf-8") as f:
            json.dump(self.templates, f, ensure_ascii=False, indent=2)

    @staticmethod
    def _header_similarity(headers_a: List[str], headers_b: List[str]) -> float:
        """计算两个表头列表的相似度（0~1）。
        用 set 交集 / 并集 的 Jaccard 相似度 + 长度加权。
        """
        if not headers_a and not headers_b:
            return 1.0
        if not headers_a or not headers_b:
            return 0.0
        set_a = {str(h).strip().lower() for h in headers_a}
        set_b = {str(h).strip().lower() for h in headers_b}
        intersection = len(set_a & set_b)
        union = len(set_a | set_b)
        if union == 0:
            return 0.0
        return intersection / union

    def find_best_match(self, headers: List[str]) -> Tuple[Optional[Dict[str, Any]], float]:
        """找到与当前表头最匹配的模板，返回 (模板, 相似度)。"""
        best_template = None
        best_score = 0.0
        for tpl in self.templates:
            score = self._header_similarity(headers, tpl.get("headers", []))
            if score > best_score:
                best_score = score
                best_template = tpl
        return best_template, best_score

    def list_templates(self) -> List[Dict[str, Any]]:
        """列出所有模板（按最后使用时间倒序）。"""
        return sorted(
            self.templates,
            key=lambda t: t.get("last_used_at", ""),
            reverse=True
        )

    def save_template(self, store_name: str, headers: List[str],
                      mapping: Dict[str, int]) -> Dict[str, Any]:
        """保存或更新模板。如果同名则覆盖。"""
        now = datetime.now().isoformat()
        existing_idx = next(
            (i for i, t in enumerate(self.templates)
             if t.get("store_name", "").strip() == store_name.strip()),
            -1
        )
        tpl = {
            "store_name": store_name.strip(),
            "headers": list(headers),
            "mapping": dict(mapping),
            "created_at": now,
            "last_used_at": now,
            "use_count": 1,
        }
        if existing_idx >= 0:
            tpl["created_at"] = self.templates[existing_idx].get("created_at", now)
            tpl["use_count"] = self.templates[existing_idx].get("use_count", 0) + 1
            self.templates[existing_idx] = tpl
        else:
            self.templates.append(tpl)
        self._save()
        return tpl

    def touch_template(self, store_name: str):
        """标记模板为已使用（更新最后使用时间和次数）。"""
        for tpl in self.templates:
            if tpl.get("store_name", "") == store_name:
                tpl["last_used_at"] = datetime.now().isoformat()
                tpl["use_count"] = tpl.get("use_count", 0) + 1
                self._save()
                break

    def delete_template(self, store_name: str) -> bool:
        before = len(self.templates)
        self.templates = [t for t in self.templates if t.get("store_name", "") != store_name]
        if len(self.templates) != before:
            self._save()
            return True
        return False


class OperationLogManager:
    """操作记录时间线管理器：记录导入、保存、批量调价、部分回滚等操作。

    每条日志结构：
    {
        "id": "uuid",
        "type": "import" | "save" | "batch_adjust" | "partial_rollback" | "full_rollback" | "snapshot",
        "action": "导入价目表",        // 人类可读操作名
        "created_at": "2025-01-01T12:00:00",
        "item_count": 120,              // 操作后项目数
        "snapshot_id": "xxx",           // 关联的快照 ID（可选）
        "detail": {                     // 操作详情，不同类型字段不同
            "source_file": "朝阳门店.xlsx",
            "added": 5, "removed": 3, "modified": 10,
            "adjust_type": "percentage", "percentage": 10.0,
            ...
        }
    }
    """

    LOG_TYPES = {
        "import": {"label": "📥 导入", "color": "#3b82f6"},
        "save": {"label": "💾 保存", "color": "#10b981"},
        "batch_adjust": {"label": "⚡ 批量调价", "color": "#f59e0b"},
        "partial_rollback": {"label": "↩️ 部分回滚", "color": "#8b5cf6"},
        "full_rollback": {"label": "⏪ 整份回滚", "color": "#ef4444"},
        "snapshot": {"label": "📸 创建快照", "color": "#0ea5e9"},
    }

    def __init__(self, data_dir: str):
        self.data_dir = data_dir
        os.makedirs(self.data_dir, exist_ok=True)
        self.log_file = os.path.join(self.data_dir, "operation_logs.json")
        self.logs: List[Dict[str, Any]] = []
        self._load()

    def _load(self):
        if os.path.exists(self.log_file):
            try:
                with open(self.log_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.logs = data if isinstance(data, list) else []
            except Exception:
                self.logs = []
        else:
            self.logs = []

    def _save(self):
        with open(self.log_file, "w", encoding="utf-8") as f:
            json.dump(self.logs, f, ensure_ascii=False, indent=2)

    def list_logs(self, limit: int = 100) -> List[Dict[str, Any]]:
        """列出所有日志（按时间倒序）。"""
        sorted_logs = sorted(self.logs, key=lambda x: x.get("created_at", ""), reverse=True)
        return sorted_logs[:limit]

    def add_log(self, log_type: str, action: str, item_count: int = 0,
                snapshot_id: str = "", detail: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """添加一条操作日志。"""
        import uuid
        log = {
            "id": str(uuid.uuid4()),
            "type": log_type,
            "action": action,
            "created_at": datetime.now().isoformat(),
            "item_count": item_count,
            "snapshot_id": snapshot_id,
            "detail": detail or {},
        }
        self.logs.insert(0, log)
        # 最多保留 500 条
        if len(self.logs) > 500:
            self.logs = self.logs[:500]
        self._save()
        return log

    def get_log(self, log_id: str) -> Optional[Dict[str, Any]]:
        return next((l for l in self.logs if l.get("id") == log_id), None)

    def clear(self):
        self.logs = []
        self._save()


class DataStore:
    def __init__(self, data_dir: str, backup_dir: str):
        self.data_dir = data_dir
        os.makedirs(self.data_dir, exist_ok=True)
        self.data_file = os.path.join(self.data_dir, "current_items.json")
        self.backup_manager = BackupManager(backup_dir)
        self.mapping_template_manager = MappingTemplateManager(data_dir)
        self.operation_log_manager = OperationLogManager(data_dir)
        self.items: List[PriceItem] = []
        self.load()

    def load(self) -> List[PriceItem]:
        self.items = []
        if os.path.exists(self.data_file):
            try:
                with open(self.data_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for item_data in data:
                    item = PriceItem(**{k: v for k, v in item_data.items()
                                       if k in PriceItem.__dataclass_fields__})
                    self.items.append(item)
            except Exception:
                self.items = []
        return self.items

    def save(self, create_snapshot: bool = True, snapshot_name: str = "",
            snapshot_desc: str = "") -> Tuple[bool, List[str]]:
        can_save, errors = PriceValidator.can_save(self.items)
        if not can_save:
            return False, errors

        if create_snapshot and self.items:
            self.backup_manager.create_snapshot(
                self.items, name=snapshot_name, description=snapshot_desc
            )

        data = [item.to_dict() for item in self.items]
        with open(self.data_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        return True, []

    def set_items(self, items: List[PriceItem]):
        self.items = items

    def add_item(self, item: PriceItem):
        self.items.append(item)

    def remove_item(self, item_id: str):
        self.items = [i for i in self.items if i.id != item_id]

    def update_item(self, item_id: str, **kwargs):
        for item in self.items:
            if item.id == item_id:
                for key, value in kwargs.items():
                    if hasattr(item, key):
                        setattr(item, key, value)
                item.update_timestamp()
                break

    def _save_data_only(self):
        data = [item.to_dict() for item in self.items]
        with open(self.data_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
